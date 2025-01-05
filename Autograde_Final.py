import os
import openpyxl

def check_bold(cell):
    """this checks if cell is bold returns bool"""
    return bool(cell.font and cell.font.bold)

def has_bottom_border(cell):
    """this checks if cell has a bottom border style"""
    return bool(cell.border and cell.border.bottom and cell.border.bottom.style)

def is_accounting_format(cell):
    """this does a rough check if a cell uses an accounting number format"""
    if not cell.number_format:
        return False
    fmt = cell.number_format
    if "$" in fmt and "0.00" in fmt:
        return True
    return False

def find_label_in_allowed_rows(sheet, label, col, allowed_rows):
    """this tries to find a label in one of the allowed rows in a given column returns the row or None"""
    label_lower = label.lower()
    for r in allowed_rows:
        val = sheet.cell(row=r, column=col).value
        if val and label_lower in str(val).lower():
            return r
    return None

def check_worksheet_labels_and_formatting(sheet, deductions):
    # we do some checks for labels offset rows and bold or not
    cell_a1 = sheet["A1"]
    if not cell_a1.value or "budget" not in str(cell_a1.value).lower():
        deductions.append(("missing or incorrect A1 label must mention budget", 5))
    else:
        if not check_bold(cell_a1):
            deductions.append(("A1 not bold", 3))

    # row 2 should be empty
    if sheet["A2"].value is not None:
        deductions.append(("row 2 A2 not empty", 2))

    # row 3 for months
    months_expected = ["Jan", "Feb", "Mar"]
    for i, month in enumerate(months_expected, start=2):
        cellval = (sheet.cell(row=3, column=i).value or "").strip()
        if cellval != month:
            deductions.append((f"month label wrong expected {month} at row 3 col {i}", 2))
        else:
            if not check_bold(sheet.cell(row=3, column=i)):
                deductions.append((f"{sheet.cell(row=3, column=i).coordinate} {month} not bold", 1))

    # row 3 also qtrly labels
    qtrly_expected = ["Qtrly Total", "Qtrly Average", "Qtrly Maximum"]
    for i, label in enumerate(qtrly_expected, start=5):
        cellval = (sheet.cell(row=3, column=i).value or "").strip()
        if cellval != label:
            deductions.append((f"qtrly label wrong expected {label} at row 3 col {i}", 2))
        else:
            if not check_bold(sheet.cell(row=3, column=i)):
                deductions.append((f"{sheet.cell(row=3, column=i).coordinate} {label} not bold", 1))

    # handle income row maybe offset
    income_row = find_label_in_allowed_rows(sheet, "Income", 1, [4,5])
    if not income_row:
        deductions.append(("cannot find Income in A4 or A5", 5))
    else:
        if not check_bold(sheet.cell(row=income_row, column=1)):
            deductions.append((f"A{income_row} Income not bold", 2))

        # next two rows are sources in non bold
        source1_row = income_row + 1
        source2_row = income_row + 2
        for r in [source1_row, source2_row]:
            if sheet.cell(row=r, column=1).value and check_bold(sheet.cell(row=r, column=1)):
                deductions.append((f"A{r} income source is bold but should not be", 1))

        # total income row and border
        total_income_row = income_row + 3
        ti_val = sheet.cell(row=total_income_row, column=1).value
        if not ti_val or "total income" not in str(ti_val).lower():
            deductions.append((f"A{total_income_row} missing total income label", 3))
        else:
            if not check_bold(sheet.cell(row=total_income_row, column=1)):
                deductions.append((f"A{total_income_row} total income not bold", 2))
        if not has_bottom_border(sheet.cell(row=total_income_row, column=1)):
            deductions.append((f"no bottom border above A{total_income_row} total income", 2))

        # empty row
        empty_row = total_income_row + 1
        if sheet.cell(row=empty_row, column=1).value:
            deductions.append((f"row {empty_row} A{empty_row} should be empty", 2))

        # expenses label row
        expenses_row = empty_row + 1
        exp_val = sheet.cell(row=expenses_row, column=1).value
        if not exp_val or "expense" not in str(exp_val).lower():
            deductions.append((f"A{expenses_row} should say Expenses", 3))
        else:
            if not check_bold(sheet.cell(row=expenses_row, column=1)):
                deductions.append((f"A{expenses_row} Expenses not bold", 2))

        # 4 expenses below non bold
        # FIX: we set exp_end = expenses_row + 5 so we get 4 lines
        exp_start = expenses_row + 1
        exp_end = expenses_row + 5  # used to be +4
        for r in range(exp_start, exp_end):
            if sheet.cell(row=r, column=1).value and check_bold(sheet.cell(row=r, column=1)):
                deductions.append((f"A{r} expense label is bold but should not be", 1))

        # total expenses row is now exp_end
        total_exp_row = exp_end
        te_val = sheet.cell(row=total_exp_row, column=1).value
        if not te_val or "total expenses" not in str(te_val).lower():
            deductions.append((f"A{total_exp_row} missing total expenses label", 3))
        else:
            if not check_bold(sheet.cell(row=total_exp_row, column=1)):
                deductions.append((f"A{total_exp_row} total expenses not bold", 2))
        if not has_bottom_border(sheet.cell(row=total_exp_row, column=1)):
            deductions.append((f"no bottom border above A{total_exp_row} total expenses", 2))

        # empty row
        empty2_row = total_exp_row + 1
        if sheet.cell(row=empty2_row, column=1).value:
            deductions.append((f"row {empty2_row} A{empty2_row} should be empty", 2))

        # net income row
        net_income_row = empty2_row + 1
        ni_val = sheet.cell(row=net_income_row, column=1).value
        if not ni_val or "net income" not in str(ni_val).lower():
            deductions.append((f"A{net_income_row} missing net income label", 3))
        else:
            if not check_bold(sheet.cell(row=net_income_row, column=1)):
                deductions.append((f"A{net_income_row} net income not bold", 2))

def check_for_addition_instead_of_range(formula_str, cell_name, deductions):
    """this looks for plus signs in formula that might mean direct additions not a range"""
    if not formula_str:
        return
    f = formula_str.upper().replace("SUM(", "").replace("AVERAGE(", "").replace("MAX(", "")
    if "+" in f:
        deductions.append((f"{cell_name} uses plus sign instead of cell range", 4))

def check_functions(sheet, deductions):
    # we loop through possible formula cells to see if they used plus signs
    for r in range(5, 21):
        for c in range(2, 8):
            cell_obj = sheet.cell(row=r, column=c)
            if cell_obj.data_type == "f":
                check_for_addition_instead_of_range(cell_obj.value, cell_obj.coordinate, deductions)

def check_accounting_format(sheet, deductions):
    # we check if at least one typical cell has an accounting format
    possible_cells = [
        "B5","C5","D5","B6","C6","D6",
        "B10","C10","D10","B11","C11","D11",
        "B12","C12","D12","B13","C13","D13"
    ]
    hits = 0
    for cell_name in possible_cells:
        if cell_name in sheet:
            cell_obj = sheet[cell_name]
            if is_accounting_format(cell_obj):
                hits += 1
                break
    if hits == 0:
        deductions.append(("no accounting format found in typical cells", 5))

def grade_excel_file(file_path):
    """this does the main checks no chart portion max 70"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active

    deductions = []

    check_worksheet_labels_and_formatting(sheet, deductions)
    check_functions(sheet, deductions)
    check_accounting_format(sheet, deductions)

    max_score = 70
    total_deduct = sum(d[1] for d in deductions)
    final_score = max_score - total_deduct
    if final_score < 0:
        final_score = 0

    return final_score, deductions

if __name__ == "__main__":
    for filename in os.listdir("."):
        if filename.lower().endswith(".xlsx"):
            score, problems = grade_excel_file(filename)
            print("")
            print(f"File: {filename}")
            print(f"Final Score ignoring chart: {score}/100")
            if not problems:
                print("no deductions perfect ignoring charts")
            else:
                print("Deductions:")
                for msg, pts in problems:
                    print(f" -{pts} pts: {msg}")
            print("-" * 60)
